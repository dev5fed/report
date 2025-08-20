import os
import streamlit as st
import pandas as pd
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


from services.db import load_planned_vs_realized_mandays


st.header("Remaining Mandays")

mapping_file = "master project mapping.xlsx"
if os.path.exists(mapping_file):
    mapping_df = pd.read_excel(mapping_file, usecols="B:C")
    mapping_df.columns = ["project_name", "project_code"]
    mapping_df = mapping_df.dropna(subset=["project_name", "project_code"])

df = load_planned_vs_realized_mandays()

# Create separate pivot tables for billable and non-billable remaining mandays
billable_pivot = df.pivot_table(
    columns="employee_code",
    index="project",
    values="remaining_billable_mandays",
    fill_value=0,
)
non_billable_pivot = df.pivot_table(
    columns="employee_code",
    index="project",
    values="remaining_non_billable_mandays",
    fill_value=0,
)

# Combine the pivot tables by adding a suffix to distinguish between billable and non-billable
billable_pivot.columns = [f"{col}_Billable" for col in billable_pivot.columns]
non_billable_pivot.columns = [
    f"{col}_NonBillable" for col in non_billable_pivot.columns
]

# Merge the two pivot tables
combined_pivot = pd.concat([billable_pivot, non_billable_pivot], axis=1, sort=True)

# Add project mapping
combined_pivot = combined_pivot.merge(
    mapping_df, left_index=True, right_on="project_code", how="left"
)
# reset index
combined_pivot = combined_pivot.reset_index()

# Reorder columns to put project information first
if "project_name" in combined_pivot.columns:
    # Get all employee codes (without suffixes) to organize columns properly
    employee_codes = sorted(
        set(
            [
                col.split("_")[0]
                for col in combined_pivot.columns
                if col.endswith(("_Billable", "_NonBillable"))
            ]
        )
    )

    # Create ordered column list: project info first, then alternating billable/non-billable for each employee
    ordered_cols = ["project_code", "project_name"]
    for emp_code in employee_codes:
        billable_col = f"{emp_code}_Billable"
        non_billable_col = f"{emp_code}_NonBillable"
        if billable_col in combined_pivot.columns:
            ordered_cols.append(billable_col)
        if non_billable_col in combined_pivot.columns:
            ordered_cols.append(non_billable_col)

    # Add any remaining columns that weren't included
    remaining_cols = [
        col
        for col in combined_pivot.columns
        if col not in ordered_cols and col != "index"
    ]
    ordered_cols.extend(remaining_cols)

    combined_pivot = combined_pivot[ordered_cols]


def create_xlsx_with_custom_headers(df):
    """
    Create an XLSX file with custom 2-row headers:
    1st row: Employee codes
    2nd row: 'Bill' and 'Non Bill' for each employee
    """
    # Create a new workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Remaining Mandays"

    # Get employee codes from the dataframe columns
    employee_codes = sorted(
        set(
            [
                col.split("_")[0]
                for col in df.columns
                if col.endswith(("_Billable", "_NonBillable"))
            ]
        )
    )

    # Create the first header row (employee codes)
    header_row1 = ["Project Code", "Project Name"]
    for emp_code in employee_codes:
        header_row1.extend([emp_code, ""])  # Employee code spans 2 columns

    # Create the second header row (Bill/Non Bill)
    header_row2 = ["", ""]  # Empty for project columns
    for emp_code in employee_codes:
        header_row2.extend(["Bill", "Non Bill"])

    # Write headers to worksheet
    ws.append(header_row1)
    ws.append(header_row2)

    # Merge cells for employee codes in first row
    col_index = 3  # Start after project columns
    for emp_code in employee_codes:
        ws.merge_cells(
            start_row=1, start_column=col_index, end_row=1, end_column=col_index + 1
        )
        col_index += 2

    # Style the headers
    for row in [1, 2]:
        for col in range(1, len(header_row1) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows
    for _, row in df.iterrows():
        data_row = [row.get("project_code", ""), row.get("project_name", "")]

        for emp_code in employee_codes:
            billable_col = f"{emp_code}_Billable"
            non_billable_col = f"{emp_code}_NonBillable"

            billable_value = row.get(billable_col, 0)
            non_billable_value = row.get(non_billable_col, 0)

            data_row.extend([billable_value, non_billable_value])

        ws.append(data_row)

    # Auto-adjust column widths

    for col_num in range(1, len(header_row1) + 1):
        max_length = 0
        column_letter = get_column_letter(col_num)

        # Check all cells in this column
        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            try:
                if hasattr(cell, "value") and cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except Exception:
                pass

        adjusted_width = min(max(max_length + 2, 10), 50)  # Minimum 10, maximum 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()


# Display the table showing both billable and non-billable remaining mandays
st.subheader("Remaining Mandays: Billable and Non-Billable")
st.dataframe(combined_pivot)

# Add download button
csv_data = combined_pivot.to_csv(index=True)
filename = f"remaining_mandays_billable_and_nonbillable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

st.download_button(
    label="📥 Download CSV",
    data=csv_data,
    file_name=filename,
    mime="text/csv",
    help="Download the remaining mandays table (both billable and non-billable) as a CSV file",
)

# Add XLSX download button with custom headers
xlsx_data = create_xlsx_with_custom_headers(combined_pivot)
xlsx_filename = f"remaining_mandays_billable_and_nonbillable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

st.download_button(
    label="📊 Download XLSX",
    data=xlsx_data,
    file_name=xlsx_filename,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    help="Download the remaining mandays table with custom headers (employee codes in 1st row, Bill/Non Bill in 2nd row)",
)
